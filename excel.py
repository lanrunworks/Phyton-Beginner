import openpyxl  # 1.
import datetime

wb = openpyxl.Workbook()  # 2.
wb.get_named_ranges()  # 3.
print(wb)

sheet = wb.active  # 4.
print(sheet.title)
sheet.title = 'Sample'  # 5.

# sheet['A1'] = 'ノートPC'  # 6.
# sheet['B1'] = 3500
# sheet['C1'] = datetime.datetime.now()
sheet.cell(row=1, column=1).value = 'デスクトップpc'  # type: ignore
sheet.cell(row=1, column=2).value = 50000  # type: ignore
sheet.cell(row=1, column=3).value = datetime.date.today()  # type: ignore

shopping_list = [['モニター', 15000, datetime.date(2022, 7, 5)],
                 ['マウス', 1000, datetime.date(2022, 7, 5)]]

current_row = 1
for i, row in enumerate(shopping_list):
    for j, column in enumerate(row):
        sheet.cell(row=current_row+i+1, column=j+1).value = column

wb.save('excel_sample01.xlsx')  # 7.
